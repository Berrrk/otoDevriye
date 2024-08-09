import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd

# Geçici dosya yolları
kontrol = 'Files/kontrol.xlsx'
sablon = 'Files/sablon.xlsx'
Test = 'Files/Test.xlsx'

# kontrol.xlsx dosyasının tanımlanması ve okunması.
controlWorkbook = load_workbook(kontrol)
controlWorksheet = controlWorkbook.active

# sablon.xlsx dosyasının tanımlanması ve okunması.
templateWorkbook = load_workbook(sablon)
templateWorksheet = templateWorkbook.active

# Son satırı algılamak için pandas Index kullanımı.
lastRowIndex = pd.read_excel(kontrol)
lastRowIndex = (lastRowIndex.index[-1]) + 2 # Pandas 2 satır eksik algıladığı için 2 satır fazladan ayarlanıyor.

# Renk değişkenleri
redColor = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')
greenColor = PatternFill(start_color="00FF00", end_color="00FF00", fill_type='solid')
whiteColor = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')

# Şablon dosyasının boyanması için gerekli sütunlar
workplaceTwoColumns = ['B', 'C', 'D', 'E', 'F', 'G', 'H']

# Kontrol aşama 1 - Şablondaki tüm "+" değerine sahip hücereler kırmızıya boyanır.
def fillRedColor():
    for i in range(0, 7):
        for ii in range(2, 26):
            cell = templateWorksheet[f'{workplaceTwoColumns[i]}'+str(ii)]
            if workplaceTwoColumns[i] == 'G':
                if str(ii) == '7' or str(ii) == '8' or str(ii) == '9' or str(ii) == '10' or str(ii) == '11' or str(ii) == '12' or str(ii) == '13' or str(ii) == '14' or str(ii) == '15' or str(ii) == '16' or str(ii) == '18' or str(ii) == '19' or str(ii) == '20' or str(ii) == '24':
                    print("Fill red:", False, f"= {cell}")
                else:
                    print("Fill red:", True, f"= {cell}")
                    cell.fill = redColor
            else:
                if str(ii) == '7' or str(ii) == '8' or str(ii) == '10' or str(ii) == '12' or str(ii) == '14' or str(ii) == '16' or str(ii) == '18' or str(ii) == '19' or str(ii) == '20' or str(ii) == '24':
                    print("Fill red:", False, f"= {cell}")
                else:
                    print("Fill red:", True, f"= {cell}")
                    cell.fill = redColor

# Kontrol aşama 2 - Ana kontroller gerçekleştirilir
def control():
    daysValues=[]
    saturdayHourValues = ["01", "02", "03", "04", "05", "16", "20", "21", "22", "00"]
    nomarlHourValues = ["01", "02", "03", "04", "05", "08", "10", "12", "14", "16", "20", "21", "22", "00"]
    mondayValues = []
    tuesdayValues = []
    wednesdayValues = []
    thursdayValues = []
    fridayValues = []
    saturdayValues = []
    sundayValues = []

    startDateValue = int(input("Start Day Value: "))
    finishDateValue = int(input("Finish Day Value: "))

    monday = "0" + str(finishDateValue - 6)
    tuesday = "0" + str(finishDateValue - 5)
    wednesday = "0" + str(finishDateValue - 4)
    thursday = "0" + str(finishDateValue - 3)
    friday = "0" + str(finishDateValue - 2)
    saturday = "0" + str(finishDateValue - 1)
    sunday = "0" + str(finishDateValue)

    columnName = "B"
    columnTwoName = "D"
    startValue = 2
    finishValue = lastRowIndex

    for days in range(startDateValue, finishDateValue+1):
        daysValues.append("0"+str(days))

    for i in range(startValue, finishValue+1):
        dateCellValue = str(controlWorksheet[(columnName+str(i))].value)
        dateCellDayValue = dateCellValue[8:10]
        dateCellHourValue = dateCellValue[11:13]

        if str(dateCellDayValue) == saturday:
            if str(dateCellHourValue) in str(saturdayHourValues):
                saturdayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")
        
        elif str(dateCellDayValue) in str(daysValues):
            if str(dateCellHourValue) in str(nomarlHourValues):
                if dateCellDayValue == monday:
                    mondayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")
                if dateCellDayValue == tuesday:
                    tuesdayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")
                if dateCellDayValue == wednesday:
                    wednesdayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")  
                if dateCellDayValue == thursday:
                    thursdayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")
                if dateCellDayValue == friday:
                    fridayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")
                if dateCellDayValue == sunday:
                    sundayValues.append(f"{dateCellHourValue} : {controlWorksheet[(columnTwoName+str(i))].value}")
    with open('output.txt', 'w') as dosya:
        pass
    with open('output.txt', 'a') as dosya:
        dosya.write(f"Pazartesi: {mondayValues}\n\n")
        dosya.write(f"Salı: {tuesdayValues}\n\n")
        dosya.write(f"Çarşamba: {wednesdayValues}\n\n")
        dosya.write(f"Perşembe: {thursdayValues}\n\n")
        dosya.write(f"Cuma: {fridayValues}\n\n")
        dosya.write(f"Cumartesi: {saturdayValues}\n\n")
        dosya.write(f"Pazar: {sundayValues}\n\n")

# Yapılan işlemler kaydedilir.
def save():
    templateFileName = 'rapor.xlsx' # Django'ya geçince değiştirilecek.
    controlFileName = 'kontrolson.xlsx' # Django'ya geçince değiştirilecek.
    templateWorkbook.save(templateFileName)
    controlWorkbook.save(controlFileName)

fillRedColor()
control()
save()